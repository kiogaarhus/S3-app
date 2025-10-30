"""
Reports API endpoints for Analytics and Business Intelligence.
Provides comprehensive reporting with caching, filtering, and export capabilities.
"""

from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from io import BytesIO
import calendar

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from fastapi_cache.decorator import cache
from sqlalchemy import func, extract, case, desc, and_, or_, text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

from backend.database.session import get_db
from backend.models.amo import AMOSagsbehandling, AMOProjekt, AMOProjekttype
from backend.schemas.reports import (
    MonthlyOverviewResponse,
    ProjekttypeDistributionResponse,
    ProcessingTimesResponse,
    PaabudStatisticsResponse,
    MonthlyTrendsResponse
)

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/monthly", response_model=MonthlyOverviewResponse)
# @cache(expire=3600)  # Disabled for debugging
async def get_monthly_overview(
    year: Optional[int] = Query(None, description="Year (default: current year)"),
    month: Optional[int] = Query(None, description="Month 1-12 (default: current month)"),
    projekttype_navn: Optional[str] = Query(None, description="Filter by projekttype name"),
    db: Session = Depends(get_db)
):
    """
    Get monthly overview of cases created, completed, and active.
    
    **Parameters:**
    - year: Year to analyze (default: current year)
    - month: Month to analyze (default: current month)
    
    **Returns:**
    - created_cases: Number of cases created in the period
    - completed_cases: Number of cases completed in the period  
    - active_cases: Number of currently active cases
    - completion_rate: Percentage of cases completed
    """
    if not year:
        # Find the most recent year with data
        most_recent = db.query(
            func.max(extract('year', AMOSagsbehandling.OprettetDato))
        ).scalar()
        year = int(most_recent) if most_recent else datetime.now().year
    if not month:
        # Use a month that's likely to have data (e.g., June)
        month = 6 if year < datetime.now().year else datetime.now().month
    
    # Get start and end of month
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
    
    # Build base query with optional projekttype filter
    base_query = db.query(func.count(AMOSagsbehandling.Id))
    if projekttype_navn:
        base_query = base_query.join(
            AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id
        ).join(
            AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID
        ).filter(
            AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%")
        )
    
    # Cases created in the month (check if no data for current month, use year-to-date)
    created_cases = base_query.filter(
        AMOSagsbehandling.OprettetDato >= start_date,
        AMOSagsbehandling.OprettetDato <= end_date
    ).scalar() or 0
    
    # If no cases in current month, get year-to-date instead
    if created_cases == 0:
        year_start = datetime(year, 1, 1)
        created_cases = base_query.filter(
            AMOSagsbehandling.OprettetDato >= year_start,
            AMOSagsbehandling.OprettetDato <= datetime.now()
        ).scalar() or 0
    
    # Cases completed in the month - Use same logic as Dashboard
    # Completed = Færdigmeldt ELLER Afsluttet
    completed_cases = base_query.filter(
        or_(
            and_(
                AMOSagsbehandling.FærdigmeldingDato >= start_date,
                AMOSagsbehandling.FærdigmeldingDato <= end_date,
                AMOSagsbehandling.FærdigmeldtInt == 1
            ),
            and_(
                AMOSagsbehandling.AfsluttetDato >= start_date,
                AMOSagsbehandling.AfsluttetDato <= end_date,
                or_(
                    AMOSagsbehandling.AfsluttetInt == 1,
                    AMOSagsbehandling.AfsluttetInt == -1
                )
            )
        )
    ).scalar() or 0
    
    # If no completed cases in current month, get year-to-date instead  
    if completed_cases == 0:
        year_start = datetime(year, 1, 1)
        completed_cases = base_query.filter(
            or_(
                and_(
                    AMOSagsbehandling.FærdigmeldingDato >= year_start,
                    AMOSagsbehandling.FærdigmeldingDato <= datetime.now(),
                    AMOSagsbehandling.FærdigmeldtInt == 1
                ),
                and_(
                    AMOSagsbehandling.AfsluttetDato >= year_start,
                    AMOSagsbehandling.AfsluttetDato <= datetime.now(),
                    or_(
                        AMOSagsbehandling.AfsluttetInt == 1,
                        AMOSagsbehandling.AfsluttetInt == -1
                    )
                )
            )
        ).scalar() or 0
    
    # Currently active cases - Use same logic as Dashboard
    # Active = Hverken færdigmeldt ELLER afsluttet
    active_cases = base_query.filter(
        and_(
            # Not færdigmeldt
            or_(
                AMOSagsbehandling.FærdigmeldtInt != 1,
                AMOSagsbehandling.FærdigmeldtInt.is_(None)
            ),
            # Not afsluttet
            or_(
                AMOSagsbehandling.AfsluttetInt == 0,
                AMOSagsbehandling.AfsluttetInt.is_(None)
            )
        )
    ).scalar() or 0
    
    # Calculate completion rate
    total_cases = created_cases + completed_cases
    completion_rate = (completed_cases / total_cases * 100) if total_cases > 0 else 0
    
    return MonthlyOverviewResponse(
        success=True,
        data={
            "year": year,
            "month": month,
            "month_name": calendar.month_name[month],
            "created_cases": created_cases,
            "completed_cases": completed_cases,
            "active_cases": active_cases,
            "completion_rate": round(completion_rate, 1)
        }
    )


@router.get("/projekttyper", response_model=ProjekttypeDistributionResponse)
# @cache(expire=3600)  # Disabled for debugging
async def get_projekttype_distribution(
    from_date: Optional[datetime] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[datetime] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """
    Get distribution of cases by projekttype (Separering, Åbenland, etc.).
    
    **Parameters:**
    - from_date: Start date for filtering (default: start of year)
    - to_date: End date for filtering (default: today)
    
    **Returns:**
    - Distribution counts and percentages by projekttype
    """
    if not from_date:
        from_date = datetime(2020, 1, 1)  # Start from 2020 to capture historical data
    if not to_date:
        to_date = datetime.now()
    
    # Query projekttype distribution
    distribution = db.query(
        AMOProjekttype.ProjektType,
        func.count(AMOSagsbehandling.Id).label('count')
    ).join(
        AMOProjekt, AMOProjekt.ProjekttypeID == AMOProjekttype.ID
    ).join(
        AMOSagsbehandling, AMOSagsbehandling.ProjektID == AMOProjekt.Id
    ).filter(
        AMOSagsbehandling.OprettetDato >= from_date,
        AMOSagsbehandling.OprettetDato <= to_date
    ).group_by(
        AMOProjekttype.ProjektType
    ).order_by(
        desc(func.count(AMOSagsbehandling.Id))
    ).all()
    
    total_cases = sum(item.count for item in distribution)
    
    data = []
    for item in distribution:
        percentage = (item.count / total_cases * 100) if total_cases > 0 else 0
        data.append({
            "projekttype": item.ProjektType,
            "count": item.count,
            "percentage": round(percentage, 1)
        })
    
    return ProjekttypeDistributionResponse(
        success=True,
        data={
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "total_cases": total_cases,
            "distribution": data
        }
    )


@router.get("/sagsbehandlingstider", response_model=ProcessingTimesResponse)
# @cache(expire=3600)  # Disabled for debugging
async def get_processing_times(
    projekttype_id: Optional[int] = Query(None, description="Filter by projekttype ID"),
    db: Session = Depends(get_db)
):
    """
    Get average case processing times by projekttype.
    
    **Parameters:**
    - projekttype_id: Optional filter by specific projekttype
    
    **Returns:**
    - Average processing times overall and by projekttype
    """
    # Base query for completed cases with processing time calculation
    # Use both FærdigmeldingDato and AfsluttetDato like Dashboard logic
    # SQL Server DATEDIFF syntax: DATEDIFF(datepart, startdate, enddate)
    base_query = db.query(
        AMOProjekttype.ProjektType,
        func.avg(
            func.datediff(
                text('day'),
                AMOSagsbehandling.OprettetDato,
                func.coalesce(
                    AMOSagsbehandling.FærdigmeldingDato,
                    AMOSagsbehandling.AfsluttetDato
                )
            )
        ).label('avg_days')
    ).join(
        AMOProjekt, AMOProjekt.ProjekttypeID == AMOProjekttype.ID
    ).join(
        AMOSagsbehandling, AMOSagsbehandling.ProjektID == AMOProjekt.Id
    ).filter(
        # Case is completed (either færdigmeldt or afsluttet)
        or_(
            AMOSagsbehandling.FærdigmeldtInt == 1,
            AMOSagsbehandling.AfsluttetInt == 1,
            AMOSagsbehandling.AfsluttetInt == -1
        ),
        # Has completion date
        or_(
            AMOSagsbehandling.FærdigmeldingDato.is_not(None),
            AMOSagsbehandling.AfsluttetDato.is_not(None)
        ),
        AMOSagsbehandling.OprettetDato.is_not(None),
        # Filter to last year for better data
        AMOSagsbehandling.OprettetDato >= datetime.now() - timedelta(days=365)
    )
    
    if projekttype_id:
        base_query = base_query.filter(AMOProjekttype.ID == projekttype_id)
    
    # Get by projekttype
    by_projekttype = base_query.group_by(AMOProjekttype.ProjektType).all()
    
    # Get overall average - Use same logic as base query
    overall_avg = db.query(
        func.avg(
            func.datediff(
                text('day'),
                AMOSagsbehandling.OprettetDato,
                func.coalesce(
                    AMOSagsbehandling.FærdigmeldingDato,
                    AMOSagsbehandling.AfsluttetDato
                )
            )
        )
    ).filter(
        # Case is completed (either færdigmeldt or afsluttet)
        or_(
            AMOSagsbehandling.FærdigmeldtInt == 1,
            AMOSagsbehandling.AfsluttetInt == 1,
            AMOSagsbehandling.AfsluttetInt == -1
        ),
        # Has completion date
        or_(
            AMOSagsbehandling.FærdigmeldingDato.is_not(None),
            AMOSagsbehandling.AfsluttetDato.is_not(None)
        ),
        AMOSagsbehandling.OprettetDato.is_not(None),
        # Filter to last year for better data
        AMOSagsbehandling.OprettetDato >= datetime.now() - timedelta(days=365)
    ).scalar() or 0
    
    data = []
    for item in by_projekttype:
        avg_days = float(item.avg_days or 0)
        # Skip negative processing times (data quality issue)
        if avg_days < 0:
            print(f"WARNING: Negative processing time for {item.ProjektType}: {avg_days} days")
            avg_days = 0  # Set to 0 instead of negative
        
        data.append({
            "projekttype": item.ProjektType,
            "average_days": round(avg_days, 1)
        })
    
    return ProcessingTimesResponse(
        success=True,
        data={
            "overall_average_days": round(float(overall_avg), 1),
            "by_projekttype": data
        }
    )


@router.get("/paabud", response_model=PaabudStatisticsResponse)
@cache(expire=3600)
async def get_paabud_statistics(
    from_date: Optional[datetime] = Query(None, description="Start date"),
    to_date: Optional[datetime] = Query(None, description="End date"),
    db: Session = Depends(get_db)
):
    """
    Get påbud (enforcement order) statistics.
    
    **Parameters:**
    - from_date: Start date for filtering
    - to_date: End date for filtering
    
    **Returns:**
    - Total påbud issued
    - Active påbud
    - Overdue påbud
    - Completion statistics
    """
    if not from_date:
        from_date = datetime(2010, 1, 1)  # Start from 2010 to capture all historical påbud data
    if not to_date:
        to_date = datetime.now()
    
    # Simple påbud statistics - projekttype filter temporarily disabled
    # Total påbud cases
    total_paabud = db.query(func.count(AMOSagsbehandling.Id)).filter(
        AMOSagsbehandling.Påbud == "Ja",
        AMOSagsbehandling.OprettetDato >= from_date,
        AMOSagsbehandling.OprettetDato <= to_date
    ).scalar() or 0
    
    # Active påbud (not completed)
    active_paabud = db.query(func.count(AMOSagsbehandling.Id)).filter(
        AMOSagsbehandling.Påbud == "Ja",
        and_(
            or_(
                AMOSagsbehandling.FærdigmeldtInt != 1,
                AMOSagsbehandling.FærdigmeldtInt.is_(None)
            ),
            or_(
                AMOSagsbehandling.AfsluttetInt == 0,
                AMOSagsbehandling.AfsluttetInt.is_(None)
            )
        ),
        AMOSagsbehandling.OprettetDato >= from_date,
        AMOSagsbehandling.OprettetDato <= to_date
    ).scalar() or 0
    
    # Overdue påbud (past deadline)
    overdue_paabud = db.query(func.count(AMOSagsbehandling.Id)).filter(
        AMOSagsbehandling.Påbud == "Ja",
        AMOSagsbehandling.Påbudsfrist < datetime.now(),
        and_(
            or_(
                AMOSagsbehandling.FærdigmeldtInt != 1,
                AMOSagsbehandling.FærdigmeldtInt.is_(None)
            ),
            or_(
                AMOSagsbehandling.AfsluttetInt == 0,
                AMOSagsbehandling.AfsluttetInt.is_(None)
            )
        ),
        AMOSagsbehandling.OprettetDato >= from_date,
        AMOSagsbehandling.OprettetDato <= to_date
    ).scalar() or 0
    
    # Completed påbud
    completed_paabud = total_paabud - active_paabud
    completion_rate = (completed_paabud / total_paabud * 100) if total_paabud > 0 else 0
    
    return PaabudStatisticsResponse(
        success=True,
        data={
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "total_paabud": total_paabud,
            "active_paabud": active_paabud,
            "completed_paabud": completed_paabud,
            "overdue_paabud": overdue_paabud,
            "completion_rate": round(completion_rate, 1)
        }
    )


@router.get("/monthly-trends", response_model=MonthlyTrendsResponse)
# @cache(expire=3600)  # Disabled for debugging
async def get_monthly_trends(
    year: Optional[int] = Query(None, description="Year (default: most recent year with data)"),
    projekttype_navn: Optional[str] = Query(None, description="Filter by projekttype name"),
    db: Session = Depends(get_db)
):
    """
    Get monthly trends for cases created and completed throughout the year.
    
    **Parameters:**
    - year: Year to analyze (default: most recent year with data)
    
    **Returns:**
    - Monthly breakdown of created and completed cases
    """
    if not year:
        # Find the most recent year with data
        most_recent = db.query(
            func.max(extract('year', AMOSagsbehandling.OprettetDato))
        ).scalar()
        year = int(most_recent) if most_recent else datetime.now().year
    
    # Build base queries with optional projekttype filter
    created_query = db.query(
        extract('month', AMOSagsbehandling.OprettetDato).label('month'),
        func.count(AMOSagsbehandling.Id).label('count')
    )
    completed_query = db.query(
        extract('month', AMOSagsbehandling.FærdigmeldingDato).label('month'),
        func.count(AMOSagsbehandling.Id).label('count')
    )
    
    if projekttype_navn:
        created_query = created_query.join(
            AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id
        ).join(
            AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID
        ).filter(
            AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%")
        )
        
        completed_query = completed_query.join(
            AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id
        ).join(
            AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID
        ).filter(
            AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%")
        )
    
    # Cases created by month
    created_by_month = created_query.filter(
        extract('year', AMOSagsbehandling.OprettetDato) == year
    ).group_by(
        extract('month', AMOSagsbehandling.OprettetDato)
    ).all()
    
    # Cases completed by month
    completed_by_month = completed_query.filter(
        extract('year', AMOSagsbehandling.FærdigmeldingDato) == year,
        AMOSagsbehandling.FærdigmeldtInt == 1
    ).group_by(
        extract('month', AMOSagsbehandling.FærdigmeldingDato)
    ).all()
    
    # Convert to dictionaries for easy lookup
    created_dict = {int(item.month): item.count for item in created_by_month}
    completed_dict = {int(item.month): item.count for item in completed_by_month}
    
    # Build monthly data
    monthly_data = []
    for month in range(1, 13):
        monthly_data.append({
            "month": month,
            "month_name": calendar.month_abbr[month],
            "created": created_dict.get(month, 0),
            "completed": completed_dict.get(month, 0)
        })
    
    return MonthlyTrendsResponse(
        success=True,
        data={
            "year": year,
            "monthly_trends": monthly_data
        }
    )


@router.get("/export/excel")
async def export_reports_excel(
    report_type: str = Query(..., description="Type: monthly|projekttyper|tider|paabud"),
    from_date: Optional[datetime] = Query(None),
    to_date: Optional[datetime] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Export reports to Excel with charts and professional formatting.
    
    **Parameters:**
    - report_type: Type of report (monthly|projekttyper|tider|paabud)
    - from_date, to_date: Date range for filtering
    - year, month: For monthly reports
    
    **Returns:**
    - Excel file with data and charts
    """
    wb = Workbook()
    ws = wb.active
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    
    if report_type == "monthly":
        # Monthly overview report
        ws.title = "Monthly Overview"
        
        # Get data
        overview_data = await get_monthly_overview(year, month, db)
        data = overview_data.data
        
        # Headers
        ws['A1'] = "Metric"
        ws['B1'] = "Value"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        
        # Data
        ws['A2'] = "Year"
        ws['B2'] = data['year']
        ws['A3'] = "Month"
        ws['B3'] = data['month_name']
        ws['A4'] = "Cases Created"
        ws['B4'] = data['created_cases']
        ws['A5'] = "Cases Completed"
        ws['B5'] = data['completed_cases']
        ws['A6'] = "Active Cases"
        ws['B6'] = data['active_cases']
        ws['A7'] = "Completion Rate (%)"
        ws['B7'] = data['completion_rate']
        
    elif report_type == "projekttyper":
        # Projekttype distribution report
        ws.title = "Projekttype Distribution"
        
        # Get data
        dist_data = await get_projekttype_distribution(from_date, to_date, db)
        distribution = dist_data.data['distribution']
        
        # Headers
        ws['A1'] = "Projekttype"
        ws['B1'] = "Count"
        ws['C1'] = "Percentage"
        for col in ['A1', 'B1', 'C1']:
            ws[col].font = header_font
            ws[col].fill = header_fill
        
        # Data
        for i, item in enumerate(distribution, 2):
            ws[f'A{i}'] = item['projekttype']
            ws[f'B{i}'] = item['count']
            ws[f'C{i}'] = item['percentage']
        
        # Add pie chart
        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(distribution) + 1)
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=len(distribution) + 1)
        chart.add_data(data_ref)
        chart.set_categories(labels)
        chart.title = "Projekttype Distribution"
        ws.add_chart(chart, "E2")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f"report_{report_type}_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
