"""
API endpoints for AMOSagsbehandling (Cases).
Provides CRUD operations with filtering, pagination, export, and status management.

Task 1: API Endpoints - Sagsbehandling (Cases)
Subtasks 1.1-1.7: Schemas, CRUD, filtering, export, status workflow, tests
"""
from typing import Optional
from datetime import datetime
from io import StringIO, BytesIO
import csv

from fastapi import APIRouter, Depends, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from fastapi_cache.decorator import cache
from sqlalchemy import desc, asc, or_, cast, String, func
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from backend.database.session import get_db
from backend.models.amo import AMOSagsbehandling, AMOProjekt, AMOProjekttype
from backend.schemas.sagsbehandling import (
    SagCreate,
    SagUpdate,
    SagStatusUpdate,
    SagOut,
    SagListResponse,
    SagDetailResponse
)
from backend.utils.pagination import paginate, PaginatedResponse, PaginationMeta
from backend.utils.anlaegs_info import get_anlaegs_info_by_adresse

router = APIRouter(prefix="/api/sager", tags=["sager"])


# Debug endpoint to check data relationships
@router.get("/debug/projekt/{projekt_id}")
async def debug_projekt_data(projekt_id: int, db: Session = Depends(get_db)):
    """Debug endpoint to check projekt and import data."""
    from backend.models.amo import AMOimport

    projekt = db.query(AMOProjekt).filter(AMOProjekt.Id == projekt_id).first()

    if not projekt:
        return {"error": "Projekt not found"}

    # Check AMOimport data
    import_count = db.query(AMOimport).count()

    # Try to find match
    import_match = None
    if projekt.Projektmappe:
        import_match = db.query(AMOimport).filter(
            AMOimport.Ejendomsnummer == projekt.Projektmappe
        ).first()

    # Get sample import data
    sample_imports = db.query(AMOimport).limit(3).all()

    return {
        "projekt": {
            "Id": projekt.Id,
            "Projektnavn": projekt.Projektnavn,
            "Projektmappe": projekt.Projektmappe,
            "ProjekttypeID": projekt.ProjekttypeID
        },
        "import_stats": {
            "total_records": import_count,
            "match_found": import_match is not None,
            "match_data": {
                "Ejendomsnummer": import_match.Ejendomsnummer if import_match else None,
                "Vejnavn": import_match.Vejnavn if import_match else None,
                "Husnummer": import_match.Husnummer if import_match else None,
                "Beliggenhed": import_match.Beliggenhed if import_match else None
            } if import_match else None
        },
        "sample_imports": [
            {
                "ID": imp.ID,
                "Ejendomsnummer": imp.Ejendomsnummer,
                "Vejnavn": imp.Vejnavn,
                "Husnummer": imp.Husnummer
            } for imp in sample_imports
        ]
    }


@router.get("/debug/table-columns/{sag_id}")
async def debug_table_columns(sag_id: int, db: Session = Depends(get_db)):
    """Debug endpoint to see ALL columns in database table."""
    from sqlalchemy import text

    # Get raw row from database to see all columns
    result = db.execute(
        text("SELECT TOP 1 * FROM dbo.AMOSagsbehandling WHERE Id = :sag_id"),
        {"sag_id": sag_id}
    ).fetchone()

    if not result:
        return {"error": "Sag not found"}

    # Get column names
    columns = result._mapping.keys()

    # Return all column values
    return {
        "sag_id": sag_id,
        "available_columns": list(columns),
        "column_values": dict(result._mapping)
    }


@router.get("/debug/events/{sag_id}")
async def debug_events(sag_id: int, db: Session = Depends(get_db)):
    """Debug endpoint to check ALL events for a case, including hidden ones."""
    from backend.models.amo import AMOHændelser, AMOHændelsestyper
    from sqlalchemy import text

    # Check if case exists
    sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == sag_id).first()
    if not sag:
        return {"error": f"Sag {sag_id} not found"}

    # Get ALL events (including hidden) with raw SQL
    raw_events = db.execute(
        text("""
            SELECT h.*, ht.HændelsesType
            FROM dbo.AMOHændelser h
            LEFT JOIN dbo.AMOHændelsestyper ht ON h.TypeID = ht.Id
            WHERE h.SagsID = :sag_id
            ORDER BY h.Dato DESC
        """),
        {"sag_id": sag_id}
    ).fetchall()

    # Get count of all events
    total_count = len(raw_events) if raw_events else 0

    # Get count of visible events (Skjul != True)
    visible_events = [e for e in raw_events if not e._mapping.get('Skjul')]
    visible_count = len(visible_events)

    # Format events for inspection
    events_data = []
    for event in raw_events:
        events_data.append({
            "id": event._mapping.get('Id'),
            "sagsid": event._mapping.get('SagsID'),
            "typeid": event._mapping.get('TypeID'),
            "dato": event._mapping.get('Dato'),
            "bemaerkning": event._mapping.get('Bemærkning'),
            "init": event._mapping.get('Init'),
            "skjul": event._mapping.get('Skjul'),
            "link": event._mapping.get('Link'),
            "haendelsestype": event._mapping.get('HændelsesType')
        })

    return {
        "sag_id": sag_id,
        "total_events": total_count,
        "visible_events": visible_count,
        "hidden_events": total_count - visible_count,
        "events": events_data
    }


# Helper function for FK validation (Subtask 1.2)
def validate_projekt_exists(projekt_id: int, db: Session) -> AMOProjekt:
    """
    Validate that a projekt exists.

    Args:
        projekt_id: ID of projekt to validate
        db: Database session

    Returns:
        AMOProjekt object if found

    Raises:
        HTTPException: If projekt not found
    """
    projekt = db.query(AMOProjekt).filter(AMOProjekt.Id == projekt_id).first()
    if not projekt:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Projekt with ID {projekt_id} not found"
        )
    return projekt


def enrich_sag_with_relations(sag: AMOSagsbehandling, db: Session) -> dict:
    """
    Enrich a case with related data from joins and direct fields.

    Args:
        sag: AMOSagsbehandling instance
        db: Database session

    Returns:
        Dictionary with sag data plus joined fields and address data
    """
    sag_dict = {
        "Id": sag.Id,
        "ProjektID": sag.ProjektID,
        "Bemærkning": sag.Bemærkning,
        "OprettetDato": sag.OprettetDato,
        "AfsluttetDato": sag.AfsluttetDato,
        "Afsluttet": sag.Afsluttet,
        "AfsluttetInt": sag.AfsluttetInt,
        "Færdigmeldt": sag.Færdigmeldt,
        "FærdigmeldtInt": sag.FærdigmeldtInt,
        "FærdigmeldingDato": sag.FærdigmeldingDato,
        "AfslutUdenFærdigmelding": sag.AfslutUdenFærdigmelding,
        "Påbud": sag.Påbud,
        "Påbudsfrist": sag.Påbudsfrist,
        "projekt_navn": None,
        "projekttype_navn": None,
        # Address fields
        "ejendomsnummer": None,
        "beliggenhed": None,
        "vejnavn": None,
        "husnummer": None,
        "husbogstav": None,
        "postnummer": None,
        "by": None,
        "matrnr": None,
        "ejer": None,
        "fuld_adresse": None,
        "anlaegs_info": None,
        # Undersøgelse & Varsel
        "SkalUndersøges": sag.SkalUndersøges,
        "SkalUndersøgesDato": sag.SkalUndersøgesDato,
        "SkalUndersøgesDatoFrist": sag.SkalUndersøgesDatoFrist,
        "VarselOmPåbud": sag.VarselOmPåbud,
        "VarselDato": sag.VarselDato,
        "VarselDatoFrist": sag.VarselDatoFrist,
        # Påbud detaljer
        "Påbudsdato": sag.Påbudsdato,
        "PåbudOm": sag.PåbudOm,
        "TilladelsesDATO": sag.TilladelsesDATO,
        "KontraktDATO": sag.KontraktDATO,
        "PaabudUdloeb": sag.PaabudUdloeb,
        # Udsættelse & Indskærpelse
        "Udsættelse": sag.Udsættelse,
        "UdsættelseDato": sag.UdsættelseDato,
        "Udsættelsesfrist": sag.Udsættelsesfrist,
        "Indskærpelse": sag.Indskærpelse,
        "IndskærpelseDato": sag.IndskærpelseDato,
        "IndskærpelseFrist": sag.IndskærpelseFrist,
        # Politianmeldelse
        "Politianmeldelse": sag.Politianmeldelse,
        "PolitianmeldelseDato": sag.PolitianmeldelseDato,
        "PolitianmeldelseDatoFrist": sag.PolitianmeldelseDatoFrist,
        "PolitianmeldelseAfgjort": sag.PolitianmeldelseAfgjort,
        # Disposition & Frister
        "Disp": sag.Disp,
        "DispType": sag.DispType,
        "DispDato": sag.DispDato,
        "DispFrist": sag.DispFrist,
        "NæsteDispFrist": sag.NæsteDispFrist,
        "NæsteFristDato": sag.NæsteFristDato,
        "NæsteFristType": sag.NæsteFristType,
        # Færdigmelding & Metadata
        "RegnvandNedsives": sag.RegnvandNedsives,
        "RegistreretFærdigmeldingDato": sag.RegistreretFærdigmeldingDato,
        "Journalnummer": sag.Journalnummer,
        "SidsteRedigeretAF": sag.SidsteRedigeretAF,
        "SidstRettetDato": sag.SidstRettetDato
    }

    # Get projekt and projekttype info
    if sag.ProjektID:
        projekt = db.query(AMOProjekt).options(
            joinedload(AMOProjekt.projekttype)
        ).filter(AMOProjekt.Id == sag.ProjektID).first()

        if projekt:
            sag_dict["projekt_navn"] = projekt.Projektnavn
            if projekt.projekttype:
                sag_dict["projekttype_navn"] = projekt.projekttype.ProjektType

    # Get address and property data directly from AMOSagsbehandling table fields
    # These fields are stored directly in the sag record
    if sag.Ejendomsnummer:
        sag_dict["ejendomsnummer"] = str(sag.Ejendomsnummer)

    if sag.Adresse:
        sag_dict["beliggenhed"] = sag.Adresse
        sag_dict["fuld_adresse"] = sag.Adresse
    elif sag.Adresse2:
        sag_dict["beliggenhed"] = sag.Adresse2
        sag_dict["fuld_adresse"] = sag.Adresse2

    if sag.Postnummer:
        sag_dict["postnummer"] = str(sag.Postnummer)

    if sag.Matrnr:
        sag_dict["matrnr"] = sag.Matrnr

    if sag.EjerEtFelt:
        sag_dict["ejer"] = sag.EjerEtFelt

    # Fetch Anlægs Info from tblAdresseEjendom based on address
    # This is relevant for Åben land cases to show facility information
    if sag.Adresse:
        anlaegs_info = get_anlaegs_info_by_adresse(db, sag.Adresse)
        sag_dict["anlaegs_info"] = anlaegs_info

    return sag_dict


# Subtask 1.3: GET /sager liste med avanceret filtrering
@router.get("", response_model=PaginatedResponse[SagOut])
@cache(expire=300)  # Cache for 5 minutes
async def list_sager(
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(50, ge=1, le=1000, description="Items per page (max 1000)"),
    sort: str = Query("Id", description="Field to sort by"),
    order: str = Query("desc", description="Sort order (asc or desc)"),
    # Filtering parameters (Subtask 1.3)
    projekt_id: Optional[int] = Query(None, description="Filter by specific projekt ID"),
    projekttype_id: Optional[int] = Query(None, description="Filter by projekttype ID (via projekt)"),
    projekttype_navn: Optional[str] = Query(None, description="Filter by projekttype name (Separering, Åbenland, etc.)"),
    projekt_navn: Optional[str] = Query(None, description="Filter by projekt name"),
    faerdigmeldt: Optional[int] = Query(None, description="Filter by status (0=Active, -1=Closed, 1=Færdigmeldt only)"),
    paabud: Optional[str] = Query(None, description="Filter by påbud status"),
    oprettet_fra: Optional[datetime] = Query(None, description="Filter by creation date from (YYYY-MM-DD)"),
    oprettet_til: Optional[datetime] = Query(None, description="Filter by creation date to (YYYY-MM-DD)"),
    search: Optional[str] = Query(None, description="Search in Bemærkning field"),
    db: Session = Depends(get_db)
):
    """
    List all cases with pagination and comprehensive filtering.

    **Filtering:**
    - projekt_id: Filter by specific projekt
    - projekttype_id: Filter by projekttype via projekt relationship
    - projekttype_navn: Filter by projekttype name (e.g., 'Separering', 'Åbenland')
    - faerdigmeldt: Filter by status:
        - 0 = Active cases (not færdigmeldt AND not afsluttet)
        - -1 = Closed cases (afsluttet)
        - 1 = Færdigmeldt only
    - paabud: Filter by påbud status
    - oprettet_fra/oprettet_til: Filter by creation date range
    - search: Full-text search in Bemærkning field (case-insensitive)

    **Pagination:**
    - page: Page number (default: 1)
    - per_page: Items per page (default: 50, max: 1000)

    **Sorting:**
    - sort: Field to sort by (default: Id)
    - order: Sort order - asc or desc (default: desc for newest first)
    """
    # Build base query
    query = db.query(AMOSagsbehandling)

    # Apply filters (Subtask 1.3)
    if projekt_id is not None:
        query = query.filter(AMOSagsbehandling.ProjektID == projekt_id)

    if faerdigmeldt is not None:
        if faerdigmeldt == 0:
            # Active cases: not færdigmeldt AND not afsluttet
            # Use coalesce to handle NULL values (NULL = 0 for our comparison)
            query = query.filter(
                func.coalesce(AMOSagsbehandling.FærdigmeldtInt, 0) != 1,
                func.coalesce(AMOSagsbehandling.AfsluttetInt, 0) != 1,
                func.coalesce(AMOSagsbehandling.AfsluttetInt, 0) != -1
            )
        elif faerdigmeldt == -1:
            # Closed cases: afsluttet (regardless of færdigmeldt status)
            query = query.filter(
                or_(
                    AMOSagsbehandling.AfsluttetInt == 1,
                    AMOSagsbehandling.AfsluttetInt == -1
                )
            )
        else:
            # For other values (e.g., 1 for færdigmeldt only)
            query = query.filter(AMOSagsbehandling.FærdigmeldtInt == faerdigmeldt)

    if paabud is not None:
        query = query.filter(AMOSagsbehandling.Påbud.ilike(f"%{paabud}%"))

    if oprettet_fra is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato >= oprettet_fra)

    if oprettet_til is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato <= oprettet_til)

    if search is not None:
        # Case-insensitive search in Bemærkning
        # CAST Bemærkning (ntext) to nvarchar for LIKE operation
        query = query.filter(cast(AMOSagsbehandling.Bemærkning, String).ilike(f"%{search}%"))

    # Filter by projekttype or projekt_navn (requires join with AMOProjekt and AMOProjekttype)
    if projekttype_id is not None or projekttype_navn is not None or projekt_navn is not None:
        query = query.join(AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id)

        if projekttype_id is not None:
            query = query.filter(AMOProjekt.ProjekttypeID == projekttype_id)

        if projekttype_navn is not None:
            query = query.join(AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID)
            query = query.filter(AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%"))

        if projekt_navn is not None:
            query = query.filter(AMOProjekt.Projektnavn == projekt_navn)

    # Apply sorting
    valid_sort_fields = {"Id", "ProjektID", "OprettetDato", "AfsluttetDato", "FærdigmeldtInt", "Påbudsfrist"}
    if sort not in valid_sort_fields:
        sort = "Id"

    sort_column = getattr(AMOSagsbehandling, sort)
    if order.lower() == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(asc(sort_column))

    # Paginate
    items, pagination_meta = paginate(query, page=page, per_page=per_page)

    # Enrich with related data and convert to Pydantic models
    data = []
    for item in items:
        enriched = enrich_sag_with_relations(item, db)
        data.append(SagOut.model_validate(enriched))

    return PaginatedResponse(
        success=True,
        data=data,
        pagination=pagination_meta
    )


# Subtask 1.2: GET /sager/{id} med eager loading
@router.get("/{id}", response_model=SagDetailResponse)
@cache(expire=300)
async def get_sag(
    id: int,
    db: Session = Depends(get_db)
):
    """
    Get a single case by ID with related projekt and projekttype data.

    **Path parameters:**
    - id: Case ID

    **Returns:**
    - Case data with projekt_navn and projekttype_navn from joins
    - 404 Not Found if case doesn't exist
    """
    sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()

    if not sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    # Enrich with related data
    enriched = enrich_sag_with_relations(sag, db)

    return SagDetailResponse(
        success=True,
        data=SagOut.model_validate(enriched)
    )


# Subtask 1.2: POST /sager med FK-validering
@router.post("", response_model=SagDetailResponse, status_code=status.HTTP_201_CREATED)
async def create_sag(
    sag_data: SagCreate,
    db: Session = Depends(get_db)
):
    """
    Create a new case with FK validation.

    **Request Body:**
    ```json
    {
        "ProjektID": 1,
        "Bemærkning": "Case description",
        "Færdigmeldt": "Nej",
        "FærdigmeldtInt": 0,
        "Påbud": "Nej"
    }
    ```

    **Validation:**
    - Checks that ProjektID references an existing projekt
    - Validates required fields (ProjektID, Bemærkning)
    - Auto-sets OprettetDato to current datetime if not provided

    **Returns:**
    - 201 Created with case data
    - 404 Not Found if projekt doesn't exist
    """
    # FK validation: Check if projekt exists (Subtask 1.2)
    validate_projekt_exists(sag_data.ProjektID, db)

    # Create new sag
    db_sag = AMOSagsbehandling(
        ProjektID=sag_data.ProjektID,
        Bemærkning=sag_data.Bemærkning,
        OprettetDato=sag_data.OprettetDato or datetime.now(),
        AfsluttetDato=sag_data.AfsluttetDato,
        Afsluttet=sag_data.Afsluttet,
        AfsluttetInt=sag_data.AfsluttetInt,
        Færdigmeldt=sag_data.Færdigmeldt or "Nej",
        FærdigmeldtInt=sag_data.FærdigmeldtInt or 0,
        FærdigmeldingDato=sag_data.FærdigmeldingDato,
        AfslutUdenFærdigmelding=sag_data.AfslutUdenFærdigmelding,
        Påbud=sag_data.Påbud or "Nej",
        Påbudsfrist=sag_data.Påbudsfrist
    )

    try:
        db.add(db_sag)
        db.commit()
        db.refresh(db_sag)

        # Enrich with related data
        enriched = enrich_sag_with_relations(db_sag, db)

        return SagDetailResponse(
            success=True,
            data=SagOut.model_validate(enriched)
        )

    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Database integrity error: {str(e)}"
        )


# Subtask 1.2: PUT /sager/{id}
@router.put("/{id}", response_model=SagDetailResponse)
async def update_sag(
    id: int,
    sag_data: SagUpdate,
    db: Session = Depends(get_db)
):
    """
    Update an existing case (full update).

    **Path parameters:**
    - id: Case ID to update

    **Request Body:**
    All fields are optional. Only provided fields will be updated.

    **Validation:**
    - If ProjektID is provided, validates it exists
    """
    # Get existing sag
    db_sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()

    if not db_sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    # FK validation if ProjektID is being updated
    if sag_data.ProjektID is not None:
        validate_projekt_exists(sag_data.ProjektID, db)

    # Update fields
    update_data = sag_data.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        setattr(db_sag, field, value)

    try:
        db.commit()
        db.refresh(db_sag)

        # Enrich with related data
        enriched = enrich_sag_with_relations(db_sag, db)

        return SagDetailResponse(
            success=True,
            data=SagOut.model_validate(enriched)
        )

    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Database integrity error: {str(e)}"
        )


# Subtask 1.6: PATCH /sager/{id}/status - Status workflow endpoint
@router.patch("/{id}/status", response_model=SagDetailResponse)
async def update_sag_status(
    id: int,
    status_data: SagStatusUpdate,
    db: Session = Depends(get_db)
):
    """
    Update case status (færdigmeldt, påbud).

    **Path parameters:**
    - id: Case ID

    **Request Body:**
    ```json
    {
        "FærdigmeldtInt": 1,
        "Påbud": "Ja",
        "Påbudsfrist": "2025-12-31T00:00:00"
    }
    ```

    **Status workflow validation:**
    - FærdigmeldtInt must be 0 or 1
    - Automatically updates related text fields (Færdigmeldt, FærdigmeldingDato)

    **Returns:**
    - Updated case data
    - 404 Not Found if case doesn't exist
    """
    # Get existing sag
    db_sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()

    if not db_sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    # Update status fields
    update_data = status_data.model_dump(exclude_unset=True)

    # Status workflow logic (Subtask 1.6)
    if "FærdigmeldtInt" in update_data:
        færdigmeldt_int = update_data["FærdigmeldtInt"]

        # Validate value
        if færdigmeldt_int not in [0, 1]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="FærdigmeldtInt must be 0 or 1"
            )

        db_sag.FærdigmeldtInt = færdigmeldt_int
        db_sag.Færdigmeldt = "Ja" if færdigmeldt_int == 1 else "Nej"

        # Update færdigmelding date when marked as færdigmeldt
        if færdigmeldt_int == 1 and not db_sag.FærdigmeldingDato:
            db_sag.FærdigmeldingDato = datetime.now()

    if "Påbud" in update_data:
        db_sag.Påbud = update_data["Påbud"]

    if "Påbudsfrist" in update_data:
        db_sag.Påbudsfrist = update_data["Påbudsfrist"]

    try:
        db.commit()
        db.refresh(db_sag)

        # Enrich with related data
        enriched = enrich_sag_with_relations(db_sag, db)

        return SagDetailResponse(
            success=True,
            data=SagOut.model_validate(enriched)
        )

    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Database integrity error: {str(e)}"
        )


# Subtask 1.2: DELETE /sager/{id}
@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_sag(
    id: int,
    db: Session = Depends(get_db)
):
    """
    Delete a case by ID.

    **Path parameters:**
    - id: Case ID to delete

    **Returns:**
    - 204 No Content on success
    - 404 Not Found if case doesn't exist

    **Note:**
    This performs a hard delete. Consider implementing soft delete
    for production use.
    """
    db_sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()

    if not db_sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    try:
        db.delete(db_sag)
        db.commit()
        return None  # 204 No Content

    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete sag: {str(e)}. It may have related records."
        )


# Subtask 1.4: GET /sager/export/csv - CSV export with filtering
@router.get("/export/csv")
async def export_sager_csv(
    projekt_id: Optional[int] = Query(None),
    projekttype_id: Optional[int] = Query(None),
    projekttype_navn: Optional[str] = Query(None),
    projekt_navn: Optional[str] = Query(None),
    faerdigmeldt: Optional[int] = Query(None),
    paabud: Optional[str] = Query(None),
    oprettet_fra: Optional[datetime] = Query(None),
    oprettet_til: Optional[datetime] = Query(None),
    search: Optional[str] = Query(None),
    columns: Optional[str] = Query(None, description="Comma-separated list of columns to export"),
    db: Session = Depends(get_db)
):
    """
    Export filtered cases to CSV file.

    **Accepts all same filter parameters as GET /api/sager**

    **Column Selection:**
    - Optional 'columns' parameter (comma-separated)
    - If not provided, exports all default columns
    - Available columns: id, projekt_id, projekt_navn, projekttype_navn, bemærkning,
      oprettet_dato, afsluttet_dato, faerdigmeldt, faerdigmelding_dato, påbud, påbudsfrist, case_age, anlaegs_info

    **Returns:**
    - CSV file download with filename: sager_export_{timestamp}.csv
    - Selected columns based on 'columns' parameter or defaults

    **Formatting:**
    - Dates in YYYY-MM-DD HH:MM:SS format
    - Boolean values as 'Ja'/'Nej' (Danish)
    - UTF-8 encoding with BOM for Excel compatibility
    """

    # Define available columns and their default status
    AVAILABLE_COLUMNS = {
        'id': {'label': 'Id', 'default': True},
        'projekt_id': {'label': 'ProjektID', 'default': False},
        'projekt_navn': {'label': 'Projekt Navn', 'default': True},
        'projekttype_navn': {'label': 'Projekttype Navn', 'default': True},
        'bemærkning': {'label': 'Bemærkning', 'default': True},
        'oprettet_dato': {'label': 'OprettetDato', 'default': True},
        'afsluttet_dato': {'label': 'AfsluttetDato', 'default': False},
        'faerdigmeldt': {'label': 'Færdigmeldt', 'default': True},
        'faerdigmelding_dato': {'label': 'FærdigmeldingDato', 'default': True},
        'påbud': {'label': 'Påbud', 'default': True},
        'påbudsfrist': {'label': 'Påbudsfrist', 'default': False},
        'case_age': {'label': 'Case Age (dage)', 'default': True},

        # Metadata kolonner
        'adresse': {'label': 'Adresse', 'default': False},
        'adresse2': {'label': 'Adresse 2', 'default': False},
        'postnummer': {'label': 'Postnummer', 'default': False},
        'matrnr': {'label': 'Matrikelnummer', 'default': False},
        'ejendomsnummer': {'label': 'Ejendomsnummer', 'default': False},
        'ejer_felt': {'label': 'Ejer', 'default': False},
        'journalnummer': {'label': 'Journalnummer', 'default': False},
        'påbudsdato': {'label': 'Påbudsdato', 'default': False},
        'tilladelsesdato': {'label': 'Tilladelsesdato', 'default': False},
        'kontraktdato': {'label': 'Kontraktdato', 'default': False},
        'udsættelse': {'label': 'Udsættelse', 'default': False},
        'udsættelsedato': {'label': 'UdsættelseDato', 'default': False},
        'indskrærpelse': {'label': 'Indskærpelse', 'default': False},
        'indskrærpelsedato': {'label': 'IndskærpelseDato', 'default': False},
        'politianmeldelse': {'label': 'Politianmeldelse', 'default': False},
        'politianmeldelsedato': {'label': 'PolitianmeldelseDato', 'default': False},
        'regnvand_nedsives': {'label': 'Regnvand nedsives', 'default': False},
        'sidste_redigeret_af': {'label': 'Sidst redigeret af', 'default': False},
        'sidst_rettet_dato': {'label': 'Sidst rettet dato', 'default': False},

        # Task 13.6: Hændelser kolonner
        'haendelser_antal': {'label': 'Antal hændelser', 'default': False},
        'seneste_haendelse_dato': {'label': 'Seneste hændelse dato', 'default': False},
        'seneste_haendelse_type': {'label': 'Seneste hændelse type', 'default': False},
        'haendelser_oversigt': {'label': 'Hændelser oversigt', 'default': False},

        # BBR kolonner
        'bbr_afløbsforhold': {'label': 'BBR Afløbsforhold', 'default': False},
        'bbr_vandforsyning': {'label': 'BBR Vandforsyning', 'default': False},

        # Anlægs Info
        'anlaegs_info': {'label': 'Anlægs info', 'default': False},
    }

    # Parse selected columns
    selected_columns = []
    if columns:
        # Parse comma-separated column names
        requested_columns = [col.strip().lower() for col in columns.split(',')]
        selected_columns = [col for col in requested_columns if col in AVAILABLE_COLUMNS]
    else:
        # Use default columns
        selected_columns = [col for col, config in AVAILABLE_COLUMNS.items() if config['default']]

    # Ensure at least one column is selected
    if not selected_columns:
        selected_columns = ['id', 'projekt_navn', 'oprettet_dato']  # Minimal fallback
    # Build query with same filters as list endpoint
    query = db.query(AMOSagsbehandling)

    # Apply all filters (same logic as list_sager)
    if projekt_id is not None:
        query = query.filter(AMOSagsbehandling.ProjektID == projekt_id)

    if faerdigmeldt is not None:
        if faerdigmeldt == 0:
            # Active cases: not færdigmeldt AND not afsluttet
            # Handle NULL values explicitly (NULL != 1 returns NULL in SQL, not TRUE)
            query = query.filter(
                or_(AMOSagsbehandling.FærdigmeldtInt != 1, AMOSagsbehandling.FærdigmeldtInt.is_(None)),
                or_(AMOSagsbehandling.AfsluttetInt != 1, AMOSagsbehandling.AfsluttetInt.is_(None)),
                or_(AMOSagsbehandling.AfsluttetInt != -1, AMOSagsbehandling.AfsluttetInt.is_(None))
            )
        elif faerdigmeldt == -1:
            # Closed cases: afsluttet (regardless of færdigmeldt status)
            query = query.filter(
                or_(
                    AMOSagsbehandling.AfsluttetInt == 1,
                    AMOSagsbehandling.AfsluttetInt == -1
                )
            )
        else:
            # For other values (e.g., 1 for færdigmeldt only)
            query = query.filter(AMOSagsbehandling.FærdigmeldtInt == faerdigmeldt)

    if paabud is not None:
        query = query.filter(AMOSagsbehandling.Påbud.ilike(f"%{paabud}%"))

    if oprettet_fra is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato >= oprettet_fra)

    if oprettet_til is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato <= oprettet_til)

    if search is not None:
        # CAST Bemærkning (ntext) to nvarchar for LIKE operation
        query = query.filter(cast(AMOSagsbehandling.Bemærkning, String).ilike(f"%{search}%"))

    # Filter by projekttype or projekt_navn
    if projekttype_id is not None or projekttype_navn is not None or projekt_navn is not None:
        query = query.join(AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id)

        if projekttype_id is not None:
            query = query.filter(AMOProjekt.ProjekttypeID == projekttype_id)

        if projekttype_navn is not None:
            query = query.join(AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID)
            query = query.filter(AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%"))

        if projekt_navn is not None:
            query = query.filter(AMOProjekt.Projektnavn == projekt_navn)

    # Get all matching sager
    sager = query.all()

    # Create CSV in memory
    output = StringIO()
    # Add BOM for Excel UTF-8 compatibility
    output.write('\ufeff')

    writer = csv.writer(output)

    # Write header with selected columns
    header_row = [AVAILABLE_COLUMNS[col]['label'] for col in selected_columns]
    writer.writerow(header_row)

    # Write data rows with selected columns
    for sag in sager:
        # Get related data
        enriched = enrich_sag_with_relations(sag, db)

        # Task 13.6: Fetch hændelser for this sag if needed
        haendelser_data = None
        if any(col.startswith('haendelser_') or col.startswith('seneste_haendelse') for col in selected_columns):
            from backend.models.amo import AMOHændelser, AMOHændelsestyper
            haendelser = (
                db.query(AMOHændelser, AMOHændelsestyper.HændelsesType)
                .join(AMOHændelsestyper, AMOHændelser.TypeID == AMOHændelsestyper.Id)
                .filter(AMOHændelser.SagsID == sag.Id)
                .filter(or_(AMOHændelser.Skjul == False, AMOHændelser.Skjul.is_(None)))
                .order_by(desc(AMOHændelser.Dato))
                .all()
            )

            # Aggregate hændelser data
            haendelser_data = {
                'antal': len(haendelser),
                'seneste_dato': haendelser[0][0].Dato.strftime('%Y-%m-%d %H:%M') if haendelser and haendelser[0][0].Dato else '',
                'seneste_type': haendelser[0][1] if haendelser else '',
                'oversigt': '; '.join([
                    f"{h[1]} ({h[0].Dato.strftime('%Y-%m-%d') if h[0].Dato else 'Ingen dato'}): {(h[0].Bemærkning or '')[:50]}"
                    for h in haendelser[:5]  # Max 5 hændelser i oversigten
                ]) if haendelser else 'Ingen hændelser'
            }

        # Fetch BBR data for this sag if needed
        bbr_data = None
        if any(col.startswith('bbr_') for col in selected_columns):
            from backend.models.amo import BBRGrund, BBRGrundAfloeb
            bbr_query = db.query(BBRGrund, BBRGrundAfloeb.Beskrivelse, BBRGrundAfloeb.BurdeHaveTank)

            # Join with code table
            bbr_query = bbr_query.outerjoin(
                BBRGrundAfloeb,
                BBRGrund.Afloebsforhold == BBRGrundAfloeb.Kode
            )

            # Filter by EjendomID or AdgangsAdresseID
            if sag.Ejendomsnummer:
                bbr_query = bbr_query.filter(BBRGrund.EjendomID == sag.Ejendomsnummer)
            elif sag.AdgangsadresseID:
                bbr_query = bbr_query.filter(BBRGrund.AdgangsAdresseID == sag.AdgangsadresseID)

            bbr_result = bbr_query.first()

            if bbr_result:
                bbr_grund, beskrivelse, burde_have_tank = bbr_result
                bbr_data = {
                    'afloebsforhold': f"{bbr_grund.Afloebsforhold} - {beskrivelse}" if bbr_grund.Afloebsforhold and beskrivelse else bbr_grund.Afloebsforhold or '',
                    'vandforsyning': bbr_grund.Vandforsyning or ''
                }

        # Calculate case age
        case_age = None
        if sag.OprettetDato:
            delta = datetime.now() - sag.OprettetDato
            case_age = delta.days

        # Format dates
        oprettet_dato_str = sag.OprettetDato.strftime('%Y-%m-%d %H:%M:%S') if sag.OprettetDato else ''
        afsluttet_dato_str = sag.AfsluttetDato.strftime('%Y-%m-%d %H:%M:%S') if sag.AfsluttetDato else ''
        faerdigmelding_dato_str = sag.FærdigmeldingDato.strftime('%Y-%m-%d %H:%M:%S') if sag.FærdigmeldingDato else ''
        paabud_frist_str = sag.Påbudsfrist.strftime('%Y-%m-%d %H:%M:%S') if sag.Påbudsfrist else ''

        # Format booleans as Ja/Nej
        faerdigmeldt_str = 'Ja' if sag.FærdigmeldtInt == 1 else 'Nej'

        # Create row data based on selected columns
        row_data = []
        for col in selected_columns:
            if col == 'id':
                row_data.append(sag.Id)
            elif col == 'projekt_id':
                row_data.append(sag.ProjektID)
            elif col == 'projekt_navn':
                row_data.append(enriched.get('projekt_navn', ''))
            elif col == 'projekttype_navn':
                row_data.append(enriched.get('projekttype_navn', ''))
            elif col == 'bemærkning':
                row_data.append(sag.Bemærkning or '')
            elif col == 'oprettet_dato':
                row_data.append(oprettet_dato_str)
            elif col == 'afsluttet_dato':
                row_data.append(afsluttet_dato_str)
            elif col == 'faerdigmeldt':
                row_data.append(faerdigmeldt_str)
            elif col == 'faerdigmelding_dato':
                row_data.append(faerdigmelding_dato_str)
            elif col == 'påbud':
                row_data.append(sag.Påbud or '')
            elif col == 'påbudsfrist':
                row_data.append(paabud_frist_str)
            elif col == 'case_age':
                row_data.append(case_age or '')
            # Metadata kolonner
            elif col == 'adresse':
                row_data.append(sag.Adresse or '')
            elif col == 'adresse2':
                row_data.append(sag.Adresse2 or '')
            elif col == 'postnummer':
                row_data.append(sag.Postnummer or '')
            elif col == 'matrnr':
                row_data.append(sag.Matrnr or '')
            elif col == 'ejendomsnummer':
                row_data.append(sag.Ejendomsnummer or '')
            elif col == 'ejer_felt':
                row_data.append(sag.EjerEtFelt or '')
            elif col == 'journalnummer':
                row_data.append(sag.Journalnummer or '')
            elif col == 'påbudsdato':
                paabudsdato_str = sag.Påbudsdato.strftime('%Y-%m-%d %H:%M:%S') if sag.Påbudsdato else ''
                row_data.append(paabudsdato_str)
            elif col == 'tilladelsesdato':
                tilladelsesdato_str = sag.TilladelsesDATO.strftime('%Y-%m-%d %H:%M:%S') if sag.TilladelsesDATO else ''
                row_data.append(tilladelsesdato_str)
            elif col == 'kontraktdato':
                kontraktdato_str = sag.KontraktDATO.strftime('%Y-%m-%d %H:%M:%S') if sag.KontraktDATO else ''
                row_data.append(kontraktdato_str)
            elif col == 'udsættelse':
                row_data.append(sag.Udsættelse or '')
            elif col == 'udsættelsedato':
                udsættelsedato_str = sag.UdsættelseDato.strftime('%Y-%m-%d %H:%M:%S') if sag.UdsættelseDato else ''
                row_data.append(udsættelsedato_str)
            elif col == 'indskrærpelse':
                row_data.append(sag.Inskærpelse or '')
            elif col == 'indskrærpelsedato':
                indskrærpelsedato_str = sag.InskærpelseDato.strftime('%Y-%m-%d %H:%M:%S') if sag.InskærpelseDato else ''
                row_data.append(indskrærpelsedato_str)
            elif col == 'politianmeldelse':
                row_data.append(sag.Politianmeldelse or '')
            elif col == 'politianmeldelsedato':
                politianmeldelsedato_str = sag.PolitianmeldelseDato.strftime('%Y-%m-%d %H:%M:%S') if sag.PolitianmeldelseDato else ''
                row_data.append(politianmeldelsedato_str)
            elif col == 'regnvand_nedsives':
                row_data.append(sag.RegnvandNedsives or '')
            elif col == 'sidste_redigeret_af':
                row_data.append(sag.SidsteRedigeretAF or '')
            elif col == 'sidst_rettet_dato':
                sidst_rettet_dato_str = sag.SidstRettetDato.strftime('%Y-%m-%d %H:%M:%S') if sag.SidstRettetDato else ''
                row_data.append(sidst_rettet_dato_str)
            # Task 13.6: Hændelser kolonner
            elif col == 'haendelser_antal':
                row_data.append(haendelser_data['antal'] if haendelser_data else 0)
            elif col == 'seneste_haendelse_dato':
                row_data.append(haendelser_data['seneste_dato'] if haendelser_data else '')
            elif col == 'seneste_haendelse_type':
                row_data.append(haendelser_data['seneste_type'] if haendelser_data else '')
            elif col == 'haendelser_oversigt':
                row_data.append(haendelser_data['oversigt'] if haendelser_data else '')
            # BBR kolonner
            elif col == 'bbr_afløbsforhold':
                row_data.append(bbr_data['afloebsforhold'] if bbr_data else '')
            elif col == 'bbr_vandforsyning':
                row_data.append(bbr_data['vandforsyning'] if bbr_data else '')
            # Anlægs Info
            elif col == 'anlaegs_info':
                row_data.append(enriched.get('anlaegs_info', ''))
            else:
                row_data.append('')  # Fallback for unknown columns

        writer.writerow(row_data)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f"sager_export_{timestamp}.csv"

    # Return as streaming response
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Subtask 1.5: GET /sager/export/excel - Excel export with formatting
@router.get("/export/excel")
async def export_sager_excel(
    projekt_id: Optional[int] = Query(None),
    projekttype_id: Optional[int] = Query(None),
    projekttype_navn: Optional[str] = Query(None),
    projekt_navn: Optional[str] = Query(None),
    faerdigmeldt: Optional[int] = Query(None),
    paabud: Optional[str] = Query(None),
    oprettet_fra: Optional[datetime] = Query(None),
    oprettet_til: Optional[datetime] = Query(None),
    search: Optional[str] = Query(None),
    columns: Optional[str] = Query(None, description="Comma-separated list of columns to export"),
    db: Session = Depends(get_db)
):
    """
    Export filtered cases to professionally formatted Excel file.

    **Accepts all same filter parameters as GET /api/sager**

    **Column Selection:**
    - Optional 'columns' parameter (comma-separated)
    - If not provided, exports all default columns
    - Available columns: id, projekt_id, projekt_navn, projekttype_navn, bemærkning,
      oprettet_dato, afsluttet_dato, faerdigmeldt, faerdigmelding_dato, påbud, påbudsfrist, case_age, anlaegs_info

    **Returns:**
    - Excel file download with filename: sager_export_{timestamp}.xlsx
    - Professional formatting:
        - Bold header row with light blue background
        - Auto-filter enabled
        - Frozen header row
        - Auto-adjusted column widths
        - Borders around all cells
        - Conditional formatting for boolean columns (Ja=green, Nej=red)
        - Danish date format (DD-MM-YYYY HH:MM:SS)
    """

    # Define available columns and their default status (same as CSV)
    AVAILABLE_COLUMNS = {
        'id': {'label': 'Id', 'default': True},
        'projekt_id': {'label': 'ProjektID', 'default': False},
        'projekt_navn': {'label': 'Projekt Navn', 'default': True},
        'projekttype_navn': {'label': 'Projekttype Navn', 'default': True},
        'bemærkning': {'label': 'Bemærkning', 'default': True},
        'oprettet_dato': {'label': 'OprettetDato', 'default': True},
        'afsluttet_dato': {'label': 'AfsluttetDato', 'default': False},
        'faerdigmeldt': {'label': 'Færdigmeldt', 'default': True},
        'faerdigmelding_dato': {'label': 'FærdigmeldingDato', 'default': True},
        'påbud': {'label': 'Påbud', 'default': True},
        'påbudsfrist': {'label': 'Påbudsfrist', 'default': False},
        'case_age': {'label': 'Case Age (dage)', 'default': True},

        # Metadata kolonner
        'adresse': {'label': 'Adresse', 'default': False},
        'adresse2': {'label': 'Adresse 2', 'default': False},
        'postnummer': {'label': 'Postnummer', 'default': False},
        'matrnr': {'label': 'Matrikelnummer', 'default': False},
        'ejendomsnummer': {'label': 'Ejendomsnummer', 'default': False},
        'ejer_felt': {'label': 'Ejer', 'default': False},
        'journalnummer': {'label': 'Journalnummer', 'default': False},
        'påbudsdato': {'label': 'Påbudsdato', 'default': False},
        'tilladelsesdato': {'label': 'Tilladelsesdato', 'default': False},
        'kontraktdato': {'label': 'Kontraktdato', 'default': False},
        'udsættelse': {'label': 'Udsættelse', 'default': False},
        'udsættelsedato': {'label': 'UdsættelseDato', 'default': False},
        'indskrærpelse': {'label': 'Indskærpelse', 'default': False},
        'indskrærpelsedato': {'label': 'IndskærpelseDato', 'default': False},
        'politianmeldelse': {'label': 'Politianmeldelse', 'default': False},
        'politianmeldelsedato': {'label': 'PolitianmeldelseDato', 'default': False},
        'regnvand_nedsives': {'label': 'Regnvand nedsives', 'default': False},
        'sidste_redigeret_af': {'label': 'Sidst redigeret af', 'default': False},
        'sidst_rettet_dato': {'label': 'Sidst rettet dato', 'default': False},

        # Task 13.6: Hændelser kolonner
        'haendelser_antal': {'label': 'Antal hændelser', 'default': False},
        'seneste_haendelse_dato': {'label': 'Seneste hændelse dato', 'default': False},
        'seneste_haendelse_type': {'label': 'Seneste hændelse type', 'default': False},
        'haendelser_oversigt': {'label': 'Hændelser oversigt', 'default': False},

        # BBR kolonner
        'bbr_afløbsforhold': {'label': 'BBR Afløbsforhold', 'default': False},
        'bbr_vandforsyning': {'label': 'BBR Vandforsyning', 'default': False},

        # Anlægs Info
        'anlaegs_info': {'label': 'Anlægs info', 'default': False},
    }

    # Parse selected columns (same logic as CSV)
    selected_columns = []
    if columns:
        # Parse comma-separated column names
        requested_columns = [col.strip().lower() for col in columns.split(',')]
        selected_columns = [col for col in requested_columns if col in AVAILABLE_COLUMNS]
    else:
        # Use default columns
        selected_columns = [col for col, config in AVAILABLE_COLUMNS.items() if config['default']]

    # Ensure at least one column is selected
    if not selected_columns:
        selected_columns = ['id', 'projekt_navn', 'oprettet_dato']  # Minimal fallback
    # Build query with same filters as list endpoint (same as CSV)
    query = db.query(AMOSagsbehandling)

    # Apply all filters
    if projekt_id is not None:
        query = query.filter(AMOSagsbehandling.ProjektID == projekt_id)

    if faerdigmeldt is not None:
        if faerdigmeldt == 0:
            # Active cases: not færdigmeldt AND not afsluttet
            # Handle NULL values explicitly (NULL != 1 returns NULL in SQL, not TRUE)
            query = query.filter(
                or_(AMOSagsbehandling.FærdigmeldtInt != 1, AMOSagsbehandling.FærdigmeldtInt.is_(None)),
                or_(AMOSagsbehandling.AfsluttetInt != 1, AMOSagsbehandling.AfsluttetInt.is_(None)),
                or_(AMOSagsbehandling.AfsluttetInt != -1, AMOSagsbehandling.AfsluttetInt.is_(None))
            )
        elif faerdigmeldt == -1:
            # Closed cases: afsluttet (regardless of færdigmeldt status)
            query = query.filter(
                or_(
                    AMOSagsbehandling.AfsluttetInt == 1,
                    AMOSagsbehandling.AfsluttetInt == -1
                )
            )
        else:
            # For other values (e.g., 1 for færdigmeldt only)
            query = query.filter(AMOSagsbehandling.FærdigmeldtInt == faerdigmeldt)

    if paabud is not None:
        query = query.filter(AMOSagsbehandling.Påbud.ilike(f"%{paabud}%"))

    if oprettet_fra is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato >= oprettet_fra)

    if oprettet_til is not None:
        query = query.filter(AMOSagsbehandling.OprettetDato <= oprettet_til)

    if search is not None:
        # CAST Bemærkning (ntext) to nvarchar for LIKE operation
        query = query.filter(cast(AMOSagsbehandling.Bemærkning, String).ilike(f"%{search}%"))

    # Filter by projekttype or projekt_navn
    if projekttype_id is not None or projekttype_navn is not None or projekt_navn is not None:
        query = query.join(AMOProjekt, AMOSagsbehandling.ProjektID == AMOProjekt.Id)

        if projekttype_id is not None:
            query = query.filter(AMOProjekt.ProjekttypeID == projekttype_id)

        if projekttype_navn is not None:
            query = query.join(AMOProjekttype, AMOProjekt.ProjekttypeID == AMOProjekttype.ID)
            query = query.filter(AMOProjekttype.ProjektType.ilike(f"%{projekttype_navn}%"))

        if projekt_navn is not None:
            query = query.filter(AMOProjekt.Projektnavn == projekt_navn)

    # Get all matching sager
    sager = query.all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sager Export"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border_side = Side(style="thin", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Green fill for "Ja"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # Red fill for "Nej"
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Write header row with selected columns
    headers = [AVAILABLE_COLUMNS[col]['label'] for col in selected_columns]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_num, sag in enumerate(sager, 2):
        # Get related data
        enriched = enrich_sag_with_relations(sag, db)

        # Task 13.6: Fetch hændelser for this sag if needed
        haendelser_data = None
        if any(col.startswith('haendelser_') or col.startswith('seneste_haendelse') for col in selected_columns):
            from backend.models.amo import AMOHændelser, AMOHændelsestyper
            haendelser = (
                db.query(AMOHændelser, AMOHændelsestyper.HændelsesType)
                .join(AMOHændelsestyper, AMOHændelser.TypeID == AMOHændelsestyper.Id)
                .filter(AMOHændelser.SagsID == sag.Id)
                .filter(or_(AMOHændelser.Skjul == False, AMOHændelser.Skjul.is_(None)))
                .order_by(desc(AMOHændelser.Dato))
                .all()
            )

            # Aggregate hændelser data
            haendelser_data = {
                'antal': len(haendelser),
                'seneste_dato': haendelser[0][0].Dato.strftime('%d-%m-%Y %H:%M') if haendelser and haendelser[0][0].Dato else '',
                'seneste_type': haendelser[0][1] if haendelser else '',
                'oversigt': '; '.join([
                    f"{h[1]} ({h[0].Dato.strftime('%d-%m-%Y') if h[0].Dato else 'Ingen dato'}): {(h[0].Bemærkning or '')[:50]}"
                    for h in haendelser[:5]  # Max 5 hændelser i oversigten
                ]) if haendelser else 'Ingen hændelser'
            }

        # Fetch BBR data for this sag if needed
        bbr_data = None
        if any(col.startswith('bbr_') for col in selected_columns):
            from backend.models.amo import BBRGrund, BBRGrundAfloeb
            bbr_query = db.query(BBRGrund, BBRGrundAfloeb.Beskrivelse, BBRGrundAfloeb.BurdeHaveTank)

            # Join with code table
            bbr_query = bbr_query.outerjoin(
                BBRGrundAfloeb,
                BBRGrund.Afloebsforhold == BBRGrundAfloeb.Kode
            )

            # Filter by EjendomID or AdgangsAdresseID
            if sag.Ejendomsnummer:
                bbr_query = bbr_query.filter(BBRGrund.EjendomID == sag.Ejendomsnummer)
            elif sag.AdgangsadresseID:
                bbr_query = bbr_query.filter(BBRGrund.AdgangsAdresseID == sag.AdgangsadresseID)

            bbr_result = bbr_query.first()

            if bbr_result:
                bbr_grund, beskrivelse, burde_have_tank = bbr_result
                bbr_data = {
                    'afloebsforhold': f"{bbr_grund.Afloebsforhold} - {beskrivelse}" if bbr_grund.Afloebsforhold and beskrivelse else bbr_grund.Afloebsforhold or '',
                    'vandforsyning': bbr_grund.Vandforsyning or ''
                }

        # Calculate case age
        case_age = None
        if sag.OprettetDato:
            delta = datetime.now() - sag.OprettetDato
            case_age = delta.days

        # Format dates in Danish format
        oprettet_dato_str = sag.OprettetDato.strftime('%d-%m-%Y %H:%M:%S') if sag.OprettetDato else ''
        afsluttet_dato_str = sag.AfsluttetDato.strftime('%d-%m-%Y %H:%M:%S') if sag.AfsluttetDato else ''
        faerdigmelding_dato_str = sag.FærdigmeldingDato.strftime('%d-%m-%Y %H:%M:%S') if sag.FærdigmeldingDato else ''
        paabud_frist_str = sag.Påbudsfrist.strftime('%d-%m-%Y %H:%M:%S') if sag.Påbudsfrist else ''

        # Format færdigmeldt as Ja/Nej
        faerdigmeldt_str = 'Ja' if sag.FærdigmeldtInt == 1 else 'Nej'

        # Create row data based on selected columns (same logic as CSV)
        row_data = []
        for col in selected_columns:
            if col == 'id':
                row_data.append(sag.Id)
            elif col == 'projekt_id':
                row_data.append(sag.ProjektID)
            elif col == 'projekt_navn':
                row_data.append(enriched.get('projekt_navn', ''))
            elif col == 'projekttype_navn':
                row_data.append(enriched.get('projekttype_navn', ''))
            elif col == 'bemærkning':
                row_data.append(sag.Bemærkning or '')
            elif col == 'oprettet_dato':
                row_data.append(oprettet_dato_str)
            elif col == 'afsluttet_dato':
                row_data.append(afsluttet_dato_str)
            elif col == 'faerdigmeldt':
                row_data.append(faerdigmeldt_str)
            elif col == 'faerdigmelding_dato':
                row_data.append(faerdigmelding_dato_str)
            elif col == 'påbud':
                row_data.append(sag.Påbud or '')
            elif col == 'påbudsfrist':
                row_data.append(paabud_frist_str)
            elif col == 'case_age':
                row_data.append(case_age if case_age is not None else '')
            # Metadata kolonner
            elif col == 'adresse':
                row_data.append(sag.Adresse or '')
            elif col == 'adresse2':
                row_data.append(sag.Adresse2 or '')
            elif col == 'postnummer':
                row_data.append(sag.Postnummer or '')
            elif col == 'matrnr':
                row_data.append(sag.Matrnr or '')
            elif col == 'ejendomsnummer':
                row_data.append(sag.Ejendomsnummer or '')
            elif col == 'ejer_felt':
                row_data.append(sag.EjerEtFelt or '')
            elif col == 'journalnummer':
                row_data.append(sag.Journalnummer or '')
            elif col == 'påbudsdato':
                paabudsdato_str = sag.Påbudsdato.strftime('%d-%m-%Y %H:%M:%S') if sag.Påbudsdato else ''
                row_data.append(paabudsdato_str)
            elif col == 'tilladelsesdato':
                tilladelsesdato_str = sag.TilladelsesDATO.strftime('%d-%m-%Y %H:%M:%S') if sag.TilladelsesDATO else ''
                row_data.append(tilladelsesdato_str)
            elif col == 'kontraktdato':
                kontraktdato_str = sag.KontraktDATO.strftime('%d-%m-%Y %H:%M:%S') if sag.KontraktDATO else ''
                row_data.append(kontraktdato_str)
            elif col == 'udsættelse':
                row_data.append(sag.Udsættelse or '')
            elif col == 'udsættelsedato':
                udsættelsedato_str = sag.UdsættelseDato.strftime('%d-%m-%Y %H:%M:%S') if sag.UdsættelseDato else ''
                row_data.append(udsættelsedato_str)
            elif col == 'indskrærpelse':
                row_data.append(sag.Inskærpelse or '')
            elif col == 'indskrærpelsedato':
                indskrærpelsedato_str = sag.InskærpelseDato.strftime('%d-%m-%Y %H:%M:%S') if sag.InskærpelseDato else ''
                row_data.append(indskrærpelsedato_str)
            elif col == 'politianmeldelse':
                row_data.append(sag.Politianmeldelse or '')
            elif col == 'politianmeldelsedato':
                politianmeldelsedato_str = sag.PolitianmeldelseDato.strftime('%d-%m-%Y %H:%M:%S') if sag.PolitianmeldelseDato else ''
                row_data.append(politianmeldelsedato_str)
            elif col == 'regnvand_nedsives':
                row_data.append(sag.RegnvandNedsives or '')
            elif col == 'sidste_redigeret_af':
                row_data.append(sag.SidsteRedigeretAF or '')
            elif col == 'sidst_rettet_dato':
                sidst_rettet_dato_str = sag.SidstRettetDato.strftime('%d-%m-%Y %H:%M:%S') if sag.SidstRettetDato else ''
                row_data.append(sidst_rettet_dato_str)
            # Task 13.6: Hændelser kolonner
            elif col == 'haendelser_antal':
                row_data.append(haendelser_data['antal'] if haendelser_data else 0)
            elif col == 'seneste_haendelse_dato':
                row_data.append(haendelser_data['seneste_dato'] if haendelser_data else '')
            elif col == 'seneste_haendelse_type':
                row_data.append(haendelser_data['seneste_type'] if haendelser_data else '')
            elif col == 'haendelser_oversigt':
                row_data.append(haendelser_data['oversigt'] if haendelser_data else '')
            # BBR kolonner
            elif col == 'bbr_afløbsforhold':
                row_data.append(bbr_data['afloebsforhold'] if bbr_data else '')
            elif col == 'bbr_vandforsyning':
                row_data.append(bbr_data['vandforsyning'] if bbr_data else '')
            # Anlægs Info
            elif col == 'anlaegs_info':
                row_data.append(enriched.get('anlaegs_info', ''))
            else:
                row_data.append('')  # Fallback for unknown columns

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

            # Apply conditional formatting for Færdigmeldt column
            if col in selected_columns and col == 'faerdigmeldt':
                if value == 'Ja':
                    cell.fill = green_fill
                elif value == 'Nej':
                    cell.fill = red_fill

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f"sager_export_{timestamp}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Task 2.7: GET /sager/{id}/export/pdf - PDF export for single sag
@router.get("/{sag_id}/export/pdf")
async def export_sag_pdf(
    sag_id: int,
    db: Session = Depends(get_db)
):
    """
    Export a single case to professionally formatted PDF file.

    **Args:**
    - sag_id: ID of the case to export

    **Returns:**
    - PDF file download with filename: sag_{id}_{timestamp}.pdf
    - Professional formatting with:
        - Header with sag title and metadata
        - Organized sections (Sag Information, Status, Påbud)
        - Danish date formatting
        - Clean, readable layout
        - Automatic page breaks for long content

    **Raises:**
    - 404: If case with given ID is not found
    """
    # Fetch sag with related data
    sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == sag_id).first()

    if not sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {sag_id} not found"
        )

    # Enrich with related data
    enriched = enrich_sag_with_relations(sag, db)

    # Calculate case age
    case_age = None
    if sag.OprettetDato:
        delta = datetime.now() - sag.OprettetDato
        case_age = delta.days

    # Create PDF in memory
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    # Container for PDF elements
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=12,
        alignment=1  # Center
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=6,
        spaceBefore=12
    )
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6b7280'),
        fontName='Helvetica-Bold'
    )
    value_style = ParagraphStyle(
        'Value',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#111827')
    )

    # Header with brand (S3) on the left and AAK logo on the right
    from pathlib import Path
    assets_dir = Path(__file__).resolve().parents[2] / 'frontend' / 'public' / 'logos'
    aak_logo_file = assets_dir / 'aak-logo.png'

    brand_style = ParagraphStyle(
        'Brand', parent=styles['Heading1'], fontSize=28, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1e40af'), spaceAfter=0
    )
    brand_para = Paragraph("S3", brand_style)

    header_cells = [brand_para, ""]
    if aak_logo_file.exists():
        aak_img = Image(str(aak_logo_file))
        # Constrain logo size to fit nicely in header
        aak_img._restrictSize(5*cm, 2*cm)
        header_cells[1] = aak_img

    header_table = Table([header_cells], colWidths=[12*cm, 5*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*cm))

    # Add title
    elements.append(Paragraph(f"Sag #{sag.Id}", title_style))
    elements.append(Spacer(1, 0.5*cm))

    # Add projekt info if available
    if enriched.get('projekt_navn'):
        elements.append(Paragraph(f"Projekt: {enriched['projekt_navn']}", value_style))
    if enriched.get('projekttype_navn'):
        projekttype_badge = f"<b>Projekttype:</b> {enriched['projekttype_navn']}"
        elements.append(Paragraph(projekttype_badge, value_style))
    elements.append(Spacer(1, 0.5*cm))

    # Status badges
    status_text = "✅ Færdigmeldt" if sag.FærdigmeldtInt == 1 else "🔵 Aktiv"
    elements.append(Paragraph(f"<b>Status:</b> {status_text}", value_style))
    if case_age is not None:
        age_indicator = " ⚠️" if case_age > 30 else ""
        elements.append(Paragraph(f"<b>Sagens alder:</b> {case_age} dage{age_indicator}", value_style))
    elements.append(Spacer(1, 0.8*cm))

    # Sag Information Section
    elements.append(Paragraph("📋 Sag Information", heading_style))

    sag_info_data = []
    sag_info_data.append(['Projekt ID', str(sag.ProjektID)])
    sag_info_data.append(['Projekt Navn', enriched.get('projekt_navn', '-')])

    if sag.OprettetDato:
        oprettet_str = sag.OprettetDato.strftime('%d-%m-%Y %H:%M')
        sag_info_data.append(['Oprettet Dato', oprettet_str])

    if sag.Bemærkning:
        # Use Paragraph for text wrapping in long bemærkninger
        bemarkung_para = Paragraph(str(sag.Bemærkning), value_style)
        sag_info_data.append(['Bemærkning', bemarkung_para])

    sag_info_table = Table(sag_info_data, colWidths=[5*cm, 12*cm])
    sag_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Top align for wrapped text
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    elements.append(sag_info_table)
    elements.append(Spacer(1, 0.5*cm))

    # Status Information Section
    elements.append(Paragraph("📊 Status Information", heading_style))

    status_info_data = []
    status_info_data.append(['Færdigmeldt', 'Ja' if sag.FærdigmeldtInt == 1 else 'Nej'])

    if sag.FærdigmeldingDato:
        faerdig_str = sag.FærdigmeldingDato.strftime('%d-%m-%Y')
        status_info_data.append(['Færdigmelding Dato', faerdig_str])

    status_info_data.append(['Afsluttet', 'Ja' if sag.AfsluttetInt == 1 else 'Nej'])

    if sag.AfsluttetDato:
        afsluttet_str = sag.AfsluttetDato.strftime('%d-%m-%Y')
        status_info_data.append(['Afsluttet Dato', afsluttet_str])

    status_info_table = Table(status_info_data, colWidths=[5*cm, 12*cm])
    status_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    elements.append(status_info_table)
    elements.append(Spacer(1, 0.5*cm))

    # Påbud Information (if applicable)
    if sag.Påbud or sag.Påbudsfrist:
        elements.append(Paragraph("⚠️ Påbud Information", heading_style))

        paabud_data = []
        if sag.Påbud:
            paabud_data.append(['Påbud', str(sag.Påbud)])

        if sag.Påbudsfrist:
            paabud_frist_str = sag.Påbudsfrist.strftime('%d-%m-%Y')
            paabud_data.append(['Påbudsfrist', paabud_frist_str])

        paabud_table = Table(paabud_data, colWidths=[5*cm, 12*cm])
        paabud_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fffbeb')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#92400e')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#78350f')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#f59e0b'))
        ]))
        elements.append(paabud_table)
        elements.append(Spacer(1, 0.5*cm))

    # Anlægs Info Section (if applicable)
    if enriched.get('anlaegs_info') and enriched.get('anlaegs_info') != "ingen bemærkninger":
        elements.append(Paragraph("🏗️ Anlægs info", heading_style))

        anlaegs_para = Paragraph(enriched['anlaegs_info'], value_style)
        anlaegs_data = [['Anlægs info', anlaegs_para]]

        anlaegs_table = Table(anlaegs_data, colWidths=[5*cm, 12*cm])
        anlaegs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eff6ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
        ]))
        elements.append(anlaegs_table)
        elements.append(Spacer(1, 0.5*cm))

    # Task 13.7: Hændelser Section
    from backend.models.amo import AMOHændelser, AMOHændelsestyper
    haendelser = (
        db.query(AMOHændelser, AMOHændelsestyper.HændelsesType)
        .join(AMOHændelsestyper, AMOHændelser.TypeID == AMOHændelsestyper.Id)
        .filter(AMOHændelser.SagsID == sag_id)
        .filter(or_(AMOHændelser.Skjul == False, AMOHændelser.Skjul.is_(None)))
        .order_by(desc(AMOHændelser.Dato))
        .all()
    )

    if haendelser:
        elements.append(Paragraph("📅 Hændelseshistorik", heading_style))

        # Build table data
        haendelser_data = [['Dato', 'Type', 'Bemærkning', 'Init.']]  # Header row

        for haendelse, type_navn in haendelser:
            dato_str = haendelse.Dato.strftime('%d-%m-%Y %H:%M') if haendelse.Dato else '-'

            # Wrap long bemærkninger
            bemaerkning_text = (haendelse.Bemærkning or '-')[:100]  # Limit length
            bemaerkning_para = Paragraph(bemaerkning_text, ParagraphStyle(
                'HaendelseBemaerkning',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#374151')
            ))

            init_str = haendelse.Init or '-'

            haendelser_data.append([
                dato_str,
                type_navn or '-',
                bemaerkning_para,
                init_str
            ])

        # Create table with appropriate column widths
        haendelser_table = Table(haendelser_data, colWidths=[3.5*cm, 3.5*cm, 7.5*cm, 2.5*cm])
        haendelser_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#374151')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Date left aligned
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Type left aligned
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Bemærkning left aligned
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Init centered
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),

            # Borders and padding
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        elements.append(haendelser_table)
    else:
        elements.append(Paragraph("📅 Hændelseshistorik", heading_style))
        elements.append(Paragraph("<i>Ingen hændelser registreret</i>", ParagraphStyle(
            'NoHaendelser',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#9ca3af'),
            alignment=1  # Center
        )))

    # Add footer with generation timestamp
    elements.append(Spacer(1, 1*cm))
    footer_text = f"<i>Genereret: {datetime.now().strftime('%d-%m-%Y kl. %H:%M')}</i>"
    elements.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#9ca3af'),
        alignment=1  # Center
    )))

    # Build PDF
    doc.build(elements)
    output.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    filename = f"sag_{sag_id}_{timestamp}.pdf"

    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Task 13.1: GET /sager/{id}/haendelser - Hent hændelser for en sag
@router.get("/{id}/haendelser", response_model=dict)
@cache(expire=300)  # Cache for 5 minutes
async def get_sag_haendelser(
    id: int,
    db: Session = Depends(get_db)
):
    """
    Get all events (hændelser) for a specific case.

    **Path parameters:**
    - id: Case ID

    **Returns:**
    - List of events with type information
    - Events are sorted by date (newest first)
    - Hidden events (Skjul=true) are excluded
    - 404 Not Found if case doesn't exist

    **Response format:**
    ```json
    {
        "success": true,
        "data": [
            {
                "id": 1,
                "dato": "2025-01-15T10:30:00",
                "haendelsestype": "Påbud",
                "bemaerkning": "Påbud udstedt",
                "init": "ABC",
                "link": null
            }
        ],
        "count": 1
    }
    ```
    """
    from backend.models.amo import AMOHændelser, AMOHændelsestyper

    # Verify that the case exists
    sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()
    if not sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    # Query hændelser with LEFT JOIN to get type name (includes events with TypeID = null)
    haendelser = (
        db.query(AMOHændelser, AMOHændelsestyper.HændelsesType)
        .outerjoin(AMOHændelsestyper, AMOHændelser.TypeID == AMOHændelsestyper.Id)
        .filter(AMOHændelser.SagsID == id)
        .filter(or_(AMOHændelser.Skjul == False, AMOHændelser.Skjul.is_(None)))
        .order_by(desc(AMOHændelser.Dato))
        .all()
    )

    # Format response
    data = []
    for haendelse, type_navn in haendelser:
        data.append({
            "id": haendelse.Id,
            "dato": haendelse.Dato,
            "haendelsestype": type_navn or "Uden type",  # Provide default for null types
            "bemaerkning": haendelse.Bemærkning,
            "init": haendelse.Init,
            "link": haendelse.Link
        })

    return {
        "success": True,
        "data": data,
        "count": len(data)
    }


# Task 14.1: GET /sager/{id}/registreringer - Hent sag-registreringer (BBR status, etc.)
@router.get("/{id}/registreringer")
@cache(expire=300)  # Cache for 5 minutes
async def get_sag_registreringer(
    id: int,
    db: Session = Depends(get_db)
):
    """
    Get all registrations (dynamiske felter) for a specific case.

    These are dynamic fields like BBR status that are stored in AMOSagReg table.

    **Path parameters:**
    - id: Case ID

    **Returns:**
    - List of registrations with field names from AMOProjekttypeReg
    - 404 Not Found if case doesn't exist

    **Response format:**
    ```json
    {
        "success": true,
        "data": [
            {
                "id": 1,
                "felt_navn": "BBR status",
                "værdi": "502",
                "dato": "2025-01-15T10:30:00",
                "frist": null,
                "init": "ABC"
            }
        ],
        "count": 1
    }
    ```
    """
    from backend.models.amo import AMOSagReg, AMOProjekttypeReg, BBRGrund, BBRGrundAfloeb

    # Verify that the case exists
    sag = db.query(AMOSagsbehandling).filter(AMOSagsbehandling.Id == id).first()
    if not sag:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sag with ID {id} not found"
        )

    # Query registrations with join to get field names
    registreringer = (
        db.query(AMOSagReg, AMOProjekttypeReg.Navn)
        .join(AMOProjekttypeReg, AMOSagReg.PTregID == AMOProjekttypeReg.ID)
        .filter(AMOSagReg.SagID == id)
        .order_by(AMOProjekttypeReg.Navn)
        .all()
    )

    # Format response
    data = []
    for reg, felt_navn in registreringer:
        data.append({
            "id": reg.ID,
            "felt_navn": felt_navn,
            "værdi": reg.JaNej,
            "dato": reg.Dato,
            "frist": reg.Frist,
            "init": reg.Init
        })

    # Task 14: Fetch BBR data if available
    if sag.Ejendomsnummer or sag.AdgangsadresseID:
        bbr_query = db.query(BBRGrund, BBRGrundAfloeb.Beskrivelse, BBRGrundAfloeb.BurdeHaveTank)

        # Join with code table
        bbr_query = bbr_query.outerjoin(
            BBRGrundAfloeb,
            BBRGrund.Afloebsforhold == BBRGrundAfloeb.Kode
        )

        # Filter by EjendomID or AdgangsAdresseID
        if sag.Ejendomsnummer:
            bbr_query = bbr_query.filter(BBRGrund.EjendomID == sag.Ejendomsnummer)
        elif sag.AdgangsadresseID:
            bbr_query = bbr_query.filter(BBRGrund.AdgangsAdresseID == sag.AdgangsadresseID)

        bbr_data = bbr_query.first()

        if bbr_data:
            bbr_grund, beskrivelse, burde_have_tank = bbr_data

            # Add BBR Afloebsforhold
            if bbr_grund.Afloebsforhold:
                værdi_text = f"{bbr_grund.Afloebsforhold}"
                if beskrivelse:
                    værdi_text += f" - {beskrivelse}"

                data.append({
                    "id": f"bbr_{bbr_grund.ID}",
                    "felt_navn": "BBR Afløbsforhold",
                    "værdi": værdi_text,
                    "dato": None,
                    "frist": None,
                    "init": None
                })

            # Add Vandforsyning if available
            if bbr_grund.Vandforsyning:
                data.append({
                    "id": f"bbr_vand_{bbr_grund.ID}",
                    "felt_navn": "BBR Vandforsyning",
                    "værdi": bbr_grund.Vandforsyning,
                    "dato": None,
                    "frist": None,
                    "init": None
                })

    return {
        "success": True,
        "data": data,
        "count": len(data)
    }
